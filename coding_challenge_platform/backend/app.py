import requests
# --- Helper for Judge0 API (Free Compiler API) ---
def judge0_compile(source_code, language_id, stdin=None):
    """
    Compile and run code using Judge0 API. Returns output, stderr, and status.
    language_id: see https://ce.judge0.com/languages (e.g., 71 for Python 3)
    """
    url = "https://ce.judge0.com/submissions/?base64_encoded=false&wait=true"
    payload = {
        "source_code": source_code,
        "language_id": language_id,
        "stdin": stdin or ""
    }
    try:
        resp = requests.post(url, json=payload, timeout=10)
        if resp.status_code == 201:
            result = resp.json()
            return {
                "stdout": result.get("stdout", ""),
                "stderr": result.get("stderr", ""),
                "status": result.get("status", {}).get("description", ""),
                "time": result.get("time", "")
            }
        else:
            return {"error": f"Judge0 error: {resp.status_code}"}
    except Exception as e:
        return {"error": str(e)}

# --- Round 3: Check Code (compile/run before submit) ---
@app.route('/check_debug_code', methods=['POST'])
def check_debug_code():
    data = request.get_json()
    code = data.get('code')
    language = data.get('lang', 'py')
    stdin = data.get('input', '')
    # Map language to Judge0 language_id
    lang_map = {'py': 71, 'c': 50, 'cpp': 54, 'java': 62}
    language_id = lang_map.get(language, 71)
    result = judge0_compile(code, language_id, stdin)
    return jsonify(result)
import os
import random
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
import subprocess
import difflib
from openpyxl import load_workbook
from datetime import datetime
import secrets
from pymongo import MongoClient
from bson import ObjectId

app = Flask(__name__, static_folder='../frontend', static_url_path='')
CORS(app)  # Enable CORS for all routes

# Root route - serve the main page
@app.route("/")
def index():
    return send_from_directory('../frontend', 'index.html')

# Serve static files (CSS, JS, images)
@app.route("/<path:path>")
def serve_static(path):
    return send_from_directory('../frontend', path)

@app.route("/api/hello")
def hello():
    return {"message":"Hello from backend"}

# --- Configuration & Initialization ---
UPLOAD_FOLDER = 'uploads'
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "adminpass"

# --- MongoDB Configuration ---
# Use a remote MongoDB instance
MONGO_URI = os.environ.get("MONGO_URI")
DB_NAME = "coding_challenge"
SCORES_COLLECTION_NAME = "scores"

try:
    mongo_client = MongoClient(MONGO_URI)
    mongo_db = mongo_client[DB_NAME]
    scores_collection = mongo_db[SCORES_COLLECTION_NAME]
    users_collection = mongo_db["users"]
    print("Successfully connected to MongoDB.")
except Exception as e:
    print(f"Error connecting to MongoDB: {e}")
    scores_collection = None
    users_collection = None

def init_db():
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'mcq'), exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'scramble', 'py'), exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'scramble', 'c'), exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'scramble', 'cpp'), exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'scramble', 'java'), exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'debug', 'py'), exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'debug', 'c'), exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'debug', 'cpp'), exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'debug', 'java'), exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'frontend_submissions'), exist_ok=True)

# --- Login Endpoints ---
@app.route('/admin_login', methods=['POST'])
def admin_login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    # Check for hardcoded admin
    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        return jsonify({"message": "Admin login successful!"}), 200
    
    # Check users collection
    if users_collection is not None:
        user = users_collection.find_one({"username": username, "role": "admin"})
        if user and user.get("password") == password:
            return jsonify({"message": "Admin login successful!"}), 200
    
    return jsonify({"message": "Invalid credentials"}), 401

@app.route('/student_login', methods=['POST'])
def student_login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({"message": "Username and password are required"}), 400
    
    # Check users collection
    if users_collection is not None:
        user = users_collection.find_one({"username": username, "role": "student"})
        if user and user.get("password") == password:
            return jsonify({"message": "Student login successful!"}), 200
        elif user:
            return jsonify({"message": "Invalid password"}), 401
        else:
            return jsonify({"message": "Username not found. Please sign up first."}), 404
    
    return jsonify({"message": "Login service unavailable"}), 500

@app.route('/student_signup', methods=['POST'])
def student_signup():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({"message": "Username and password are required"}), 400
    
    if users_collection is not None:
        # Check if user already exists
        existing_user = users_collection.find_one({"username": username, "role": "student"})
        if existing_user:
            return jsonify({"message": "Team username already exists. Please choose another."}), 400
        
        # Create new student user
        try:
            users_collection.insert_one({
                "username": username,
                "password": password,
                "role": "student",
                "created_at": datetime.now()
            })
            return jsonify({"message": "Team account created successfully!"}), 201
        except Exception as e:
            return jsonify({"message": "Error creating account", "error": str(e)}), 500
    
    return jsonify({"message": "Sign up service unavailable"}), 500
    
# --- Admin File Upload Endpoint ---
@app.route('/admin_upload', methods=['POST'])
def admin_upload():
    round_name = request.form.get('round')
    lang = request.form.get('lang')
    file = request.files.get('file')

    if not file or file.filename == '':
        return jsonify({"message": "No file part"}), 400

    if round_name == 'mcq':
        if not file.filename.endswith('.xlsx'):
            return jsonify({"message": "MCQ files must be .xlsx"}), 400
        file_path = os.path.join(UPLOAD_FOLDER, 'mcq', 'questions.xlsx')
    elif round_name == 'scramble':
        if lang not in ['py', 'c', 'cpp', 'java']:
            return jsonify({"message": "Invalid language for scramble file."}), 400
        file_path = os.path.join(UPLOAD_FOLDER, 'scramble', lang, file.filename)
    elif round_name == 'debug':
        if lang not in ['py', 'c', 'cpp', 'java']:
            return jsonify({"message": "Invalid language for debug file."}), 400
        file_path = os.path.join(UPLOAD_FOLDER, 'debug', lang, file.filename)
    else:
        return jsonify({"message": "Invalid round selected"}), 400

    try:
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        file.save(file_path)
        return jsonify({"message": f"Successfully uploaded file for {round_name} round."}), 200
    except Exception as e:
        return jsonify({"message": "Failed to save file", "error": str(e)}), 500

# --- API Endpoints for Frontend Interaction ---
@app.route('/')
def serve_index():
    return send_from_directory(os.getcwd(), 'index.html')

# Round 1: Multiple Choice Questions
@app.route('/get_mcq_questions')
def get_mcq_questions():
    file_path = os.path.join(UPLOAD_FOLDER, 'mcq', 'questions.xlsx')
    if not os.path.exists(file_path):
        return jsonify({"error": "questions.xlsx not found. Admin needs to upload this file."}), 404
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        questions = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            question = {headers[i]: row[i].value for i in range(len(headers))}
            questions.append(question)
        random_questions = random.sample(questions, min(10, len(questions)))
        return jsonify(random_questions)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_mcq_score', methods=['POST'])
def submit_mcq_score():
    if scores_collection is None:
        return jsonify({"message": "Database not available"}), 500
    
    data = request.get_json()
    score_doc = {
        "username": data.get('username'),
        "round_name": 'MCQ',
        "correct_answers": data.get('correct_answers'),
        "total_questions": data.get('total_questions'),
        "percentage": data.get('percentage'),
        "timestamp": datetime.now()
    }
    scores_collection.insert_one(score_doc)
    
    # Return score information immediately
    return jsonify({
        "message": "Success", 
        "score": {
            "correct": data.get('correct_answers'),
            "total": data.get('total_questions'),
            "percentage": data.get('percentage')
        }
    }), 200

# Round 2: Code Scramble
@app.route('/get_scrambled_code_list', methods=['GET'])
def get_scrambled_code_list():
    lang = request.args.get('lang', 'py')
    dir_path = os.path.join(UPLOAD_FOLDER, 'scramble', lang)
    if not os.path.exists(dir_path):
        return jsonify([])
    files = [f for f in os.listdir(dir_path) if os.path.isfile(os.path.join(dir_path, f))]
    return jsonify(files)

@app.route('/get_scrambled_code', methods=['GET'])
def get_scrambled_code():
    file_path = request.args.get('file')
    if not file_path:
        return jsonify({"error": "File path not provided"}), 400

    full_path = os.path.join(UPLOAD_FOLDER, file_path)
    if not os.path.exists(full_path):
        return jsonify({"error": f"File {file_path} not found. Admin needs to upload."}), 404

    try:
        with open(full_path, 'r') as f:
            code = f.read()
            lines = [line for line in code.strip().split('\n') if line.strip()]
            random.shuffle(lines)
            scrambled_code = '\n'.join(lines)
            return jsonify({"code": scrambled_code})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_scrambled_code', methods=['POST'])
def submit_scrambled_code():
    if scores_collection is None:
        return jsonify({"message": "Database not available"}), 500
        
    data = request.get_json()
    submitted_code = data.get('code')
    file_path = data.get('file_path')
    username = data.get('username')
    lang = data.get('lang', 'py')
    
    if not file_path:
        return jsonify({"error": "File path not provided"}), 400

    full_path = os.path.join(UPLOAD_FOLDER, file_path)
    if not os.path.exists(full_path):
        return jsonify({"error": f"File {file_path} not found. Admin needs to upload."}), 404
        
    try:
        with open(full_path, 'r') as f:
            correct_code = f.read()
            correct_lines = [line.strip() for line in correct_code.strip().split('\n') if line.strip()]
    except Exception as e:
        return jsonify({"error": str(e)}), 500
        
    submitted_lines = [line.strip() for line in submitted_code.strip().split('\n') if line.strip()]
    num_attempted = len(submitted_lines)
    num_correct = sum(1 for i, line in enumerate(submitted_lines) if i < len(correct_lines) and line == correct_lines[i])
    matcher = difflib.SequenceMatcher(None, correct_lines, submitted_lines)
    score = matcher.ratio() * 100
    
    # Save submitted code to team folder structure
    if username:
        team_folder = os.path.join(UPLOAD_FOLDER, 'Round2', username, lang)
        os.makedirs(team_folder, exist_ok=True)
        
        # Create a timestamped filename for the submission
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        code_file = os.path.join(team_folder, f'submission_{timestamp}.txt')
        
        try:
            with open(code_file, 'w') as f:
                f.write(submitted_code)
        except Exception as e:
            print(f"Error saving submitted code: {e}")
        
        score_doc = {
            "username": username,
            "round_name": 'Scramble',
            "correct_answers": num_correct,
            "total_questions": num_attempted,
            "percentage": score,
            "timestamp": datetime.now()
        }
        scores_collection.insert_one(score_doc)
    
    return jsonify({
        "team_name": username,
        "attempted": num_attempted,
        "correct": num_correct,
        "accuracy": f"{score:.2f}%",
        "score": score,
        "message": "Code submitted successfully!"
    })

# Round 3: Code Debugging
@app.route('/get_buggy_code_list', methods=['GET'])
def get_buggy_code_list():
    lang = request.args.get('lang', 'py')
    dir_path = os.path.join(UPLOAD_FOLDER, 'debug', lang)
    if not os.path.exists(dir_path):
        return jsonify([])
    files = [f for f in os.listdir(dir_path) if os.path.isfile(os.path.join(dir_path, f))]
    return jsonify(files)

@app.route('/get_buggy_code', methods=['GET'])
def get_buggy_code():
    file_path = request.args.get('file')
    if not file_path:
        return jsonify({"error": "File path not provided"}), 400
    
    full_path = os.path.join(UPLOAD_FOLDER, 'debug', file_path)
    if not os.path.exists(full_path):
        return jsonify({"error": f"File {file_path} not found. Admin needs to upload."}), 404
    
    try:
        with open(full_path, 'r') as f:
            code = f.read()
            return jsonify({"code": code})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/submit_debug_code', methods=['POST'])
def submit_debug_code():
    if not scores_collection:
        return jsonify({"message": "Database not available"}), 500

    data = request.get_json()
    submitted_code = data.get('code')
    language = data.get('lang', 'py')
    file_path = data.get('file_path')
    username = data.get('username')

    if not file_path:
        return jsonify({"error": "File path not provided"}), 400

    # --- 45 min time limit logic ---
    # Find/create a start time for this team/round
    time_limit_minutes = 45
    now = datetime.now()
    start_time = None
    # Try to find the earliest submission for this team in Round3
    prev = scores_collection.find_one({"username": username, "round_name": "Debugging"}, sort=[("timestamp", 1)])
    if prev:
        start_time = prev.get("timestamp")
    else:
        start_time = now
    elapsed = (now - start_time).total_seconds() / 60.0
    if elapsed > time_limit_minutes:
        return jsonify({"message": f"Time limit exceeded ({time_limit_minutes} min). Submission not accepted."}), 403

    # --- Test cases (expand as needed) ---
    test_cases = [{'input': '10', 'expected_output': '20\n'}]
    passed_tests = 0
    total_tests = len(test_cases)
    lang_map = {'py': 71, 'c': 50, 'cpp': 54, 'java': 62}
    language_id = lang_map.get(language, 71)
    for case in test_cases:
        result = judge0_compile(submitted_code, language_id, case['input'])
        if result.get('stdout', '') == case['expected_output']:
            passed_tests += 1

    score = (passed_tests / total_tests) * 100 if total_tests else 0

    # Save submitted code to team folder structure
    if username:
        team_folder = os.path.join(UPLOAD_FOLDER, 'Round3', username, language)
        os.makedirs(team_folder, exist_ok=True)
        timestamp = now.strftime('%Y%m%d_%H%M%S')
        code_file = os.path.join(team_folder, f'submission_{timestamp}.txt')
        try:
            with open(code_file, 'w') as f:
                f.write(submitted_code)
        except Exception as e:
            print(f"Error saving submitted code: {e}")
        score_doc = {
            "username": username,
            "round_name": 'Debugging',
            "correct_answers": passed_tests,
            "total_questions": total_tests,
            "percentage": score,
            "timestamp": now,
            "time_taken_min": round(elapsed, 2)
        }
        scores_collection.insert_one(score_doc)

    return jsonify({
        "team_name": username,
        "passed_tests": passed_tests,
        "total_tests": total_tests,
        "accuracy": f"{score:.2f}%",
        "score": score,
        "time_taken_min": round(elapsed, 2),
        "message": "Code submitted and tested successfully!"
    })

@app.route('/submit_frontend', methods=['POST'])
def submit_frontend():
    if 'file' not in request.files:
        return jsonify({"message": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"message": "No selected file"}), 400
    
    # Validate file extension
    allowed_extensions = {'.txt', '.png', '.jpg', '.jpeg'}
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in allowed_extensions:
        return jsonify({"message": "Invalid file type. Only .txt, .png, and .jpg files are allowed."}), 400
    
    username = request.form.get('username', 'anonymous')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    secure_filename = f"{username}_{timestamp}_{file.filename}"
    file_path = os.path.join(UPLOAD_FOLDER, 'Round4', username, secure_filename)

    try:
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        file.save(file_path)
        return jsonify({"message": "File uploaded successfully!"}), 200
    except Exception as e:
        return jsonify({"message": "Failed to save file", "error": str(e)}), 500

# --- Admin Dashboard Endpoints ---
@app.route('/admin/scores', methods=['GET'])
def get_admin_scores():
    if scores_collection is None:
        return jsonify({"message": "Database not available"}), 500
        
    try:
        scores = list(scores_collection.find().sort("timestamp", -1))
        # Convert ObjectId and datetime to string for JSON serialization
        for score in scores:
            score['_id'] = str(score['_id'])
            score['timestamp'] = score['timestamp'].strftime('%Y-%m-%d %H:%M:%S')
        return jsonify(scores), 200
    except Exception as e:
        return jsonify({"message": "Failed to retrieve scores", "error": str(e)}), 500

# --- Student Score Viewing Endpoint ---
@app.route('/student/scores', methods=['GET'])
def get_student_scores():
    if scores_collection is None:
        return jsonify({"message": "Database not available"}), 500
    
    username = request.args.get('username')
    if not username:
        return jsonify({"message": "Username is required"}), 400
    
    try:
        # Get all scores for this student
        scores = list(scores_collection.find({"username": username}).sort("timestamp", -1))
        
        # Convert ObjectId and datetime to string for JSON serialization
        for score in scores:
            score['_id'] = str(score['_id'])
            score['timestamp'] = score['timestamp'].strftime('%Y-%m-%d %H:%M:%S')
        
        return jsonify(scores), 200
    except Exception as e:
        return jsonify({"message": "Failed to retrieve scores", "error": str(e)}), 500

@app.route('/admin/questions', methods=['GET'])
def get_admin_questions():
    file_path = os.path.join(UPLOAD_FOLDER, 'mcq', 'questions.xlsx')
    if not os.path.exists(file_path):
        return jsonify({"error": "questions.xlsx not found"}), 404
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        questions = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            question = {headers[i]: row[i].value for i in range(len(headers))}
            questions.append(question)
        return jsonify(questions)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/admin/scores/delete', methods=['DELETE'])
def delete_all_scores():
    if scores_collection is None:
        return jsonify({"message": "Database not available"}), 500
        
    try:
        scores_collection.delete_many({})
        return jsonify({"message": "All scores deleted successfully."}), 200
    except Exception as e:
        return jsonify({"message": "Failed to delete scores", "error": str(e)}), 500

@app.route('/admin/code_questions', methods=['GET'])
def get_admin_code_files():
    round_name = request.args.get('round')
    lang = request.args.get('lang')
    path = os.path.join(UPLOAD_FOLDER, round_name, lang)
    if not os.path.exists(path):
        return jsonify([])
    files = [f for f in os.listdir(path) if os.path.isfile(os.path.join(path, f))]
    return jsonify(files)

@app.route('/admin/code_content', methods=['GET'])
def get_admin_code_content():
    filename = request.args.get('filename')
    full_path = os.path.join(UPLOAD_FOLDER, filename)
    
    if not os.path.exists(full_path):
        return "File not found", 404

    try:
        with open(full_path, 'r') as f:
            content = f.read()
        return content
    except Exception as e:
        return f"Error reading file: {e}", 500

@app.route('/admin/submissions', methods=['GET'])
def get_admin_submissions():
    """Get all student submissions organized by team and round"""
    try:
        submissions = []
        rounds = {
            'Round2': 'Code Scramble',
            'Round3': 'Code Debugging',
            'Round4': 'Frontend Challenge'
        }
        
        for round_folder, round_name in rounds.items():
            round_path = os.path.join(UPLOAD_FOLDER, round_folder)
            if not os.path.exists(round_path):
                continue
                
            # List all team folders in this round
            for team_name in os.listdir(round_path):
                team_path = os.path.join(round_path, team_name)
                if not os.path.isdir(team_path):
                    continue
                
                # List all language folders or files for this team
                for item in os.listdir(team_path):
                    item_path = os.path.join(team_path, item)
                    
                    if os.path.isdir(item_path):
                        # It's a language folder (e.g., 'c', 'py', 'cpp')
                        for filename in os.listdir(item_path):
                            file_path = os.path.join(item_path, filename)
                            if os.path.isfile(file_path):
                                stat_info = os.stat(file_path)
                                submissions.append({
                                    'team_name': team_name,
                                    'round_name': round_name,
                                    'filename': filename,
                                    'language': item,
                                    'file_path': os.path.join(round_folder, team_name, item, filename),
                                    'timestamp': datetime.fromtimestamp(stat_info.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                                    'size': stat_info.st_size
                                })
                    else:
                        # It's a file directly in the team folder
                        stat_info = os.stat(item_path)
                        submissions.append({
                            'team_name': team_name,
                            'round_name': round_name,
                            'filename': item,
                            'language': 'N/A',
                            'file_path': os.path.join(round_folder, team_name, item),
                            'timestamp': datetime.fromtimestamp(stat_info.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                            'size': stat_info.st_size
                        })
        
        # Sort by timestamp (newest first)
        submissions.sort(key=lambda x: x['timestamp'], reverse=True)
        return jsonify(submissions), 200
        
    except Exception as e:
        return jsonify({"message": "Failed to retrieve submissions", "error": str(e)}), 500

@app.route('/admin/submission_content', methods=['GET'])
def get_submission_content():
    """Get the content of a specific submission file"""
    file_path = request.args.get('file_path')
    if not file_path:
        return jsonify({"error": "File path not provided"}), 400
    
    full_path = os.path.join(UPLOAD_FOLDER, file_path)
    
    if not os.path.exists(full_path):
        return jsonify({"error": "File not found"}), 404
    
    try:
        # Check if it's a text file or image
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext in ['.png', '.jpg', '.jpeg']:
            return send_from_directory(os.path.dirname(full_path), os.path.basename(full_path))
        else:
            with open(full_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            return jsonify({"content": content}), 200
    except Exception as e:
        return jsonify({"error": f"Error reading file: {e}"}), 500

if __name__ == '__main__':
    port=int(os.environ.get("Port",8000))
    init_db()
    app.run(debug=True, host='0.0.0.0', port=port)